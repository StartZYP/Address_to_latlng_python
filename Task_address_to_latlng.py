import requests
import json
from openpyxl import load_workbook

API_key ='tTDvVDh2lSg0gVb9g0NsMO8R'
file_name='excel.xlsx'

def getlat_lon(addresslist):
    lnglatlist=[]
    for i in addresslist:
        params = {
            'address': i,
            'output': 'json',
            'key': API_key
        }
        response = requests.get('http://api.map.baidu.com/geocoder', params=params)
        mytext = json.loads(response.text)
        try:
            lng = mytext['result']['location']['lng']
            lat = mytext['result']['location']['lat']
        except Exception as a:
            print(a)
            lng ='异常'
            lat ='异常'
        lnglatlist.append(str(lng)+'-'+str(lat))
    return lnglatlist

def get_excel_address(startnum,endnum):
    '''
    从三开始
    :param num:
    :return:
    '''
    if startnum<=4:
        startnum=4
    addresslist=[]
    sheet_ranges = wb['Sheet1']
    for i in range(startnum-1,endnum):
        addresslist.append(sheet_ranges['B'+str(i+1)].value)
    return addresslist

def write_excel(startnum,endnum,lnglatlist):
    if startnum<=4:
        startnum=4
    ws = wb['Sheet1']
    for i in range(startnum - 1, endnum):
        num = i-startnum+1
        lng, lat = lnglatlist[num].split('-')
        ws["I"+str(i+1)] = lng
        ws["J"+str(i+1)] = lat
    wb.save(file_name)


for i in range(1,71):
    wb = load_workbook(filename=file_name)
    num1 = i*853-853
    num2 =i*853
    print(i)
    print(str(num1)+'-'+str(num2))
    write_excel(num1, num2, getlat_lon(get_excel_address(num1, num2)))
