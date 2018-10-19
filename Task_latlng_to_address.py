#http://lbsyun.baidu.com/apiconsole/key
#申请反转地理地址API
'''
1.根据地址取得经纬度：

请求地址：http://api.map.baidu.com/geocoder/v2/?address=中国成都人才市场&output=json&ak=你的ak

2.根据经纬度取得地址：

http://api.map.baidu.com/geocoder/v2/?callback=renderReverse&location=30.68093376455154,104.06552381979525&output=json&pois=1&ak=你的ak

ak查看地址：http://lbsyun.baidu.com/apiconsole/key
'''
import requests
import json
from openpyxl import load_workbook

app_key='9elMuik91gtWCPbmW4pGdVOjplbAx5GV'
file_name='excel.xlsx'


def get_address(latlnglist):
    templist=[]
    for lat_lng in latlnglist:
        params = {
            'location':lat_lng,
            'output':'json',
            'pois':'1',
            'ak':app_key
        }
        response = requests.get('http://api.map.baidu.com/geocoder/v2/',params=params)
        jsontext = json.loads(response.text)
        try:
            templist.append(jsontext['result']['addressComponent']['province']+','+jsontext['result']['addressComponent']['city']+','+jsontext['result']['addressComponent']['district'])
        except Exception as e:
            print(e)
            templist.append('异常,异常,异常')
        # print('省:'+jsontext['result']['addressComponent']['province'])
        # print('市:'+jsontext['result']['addressComponent']['city'])
        # print('县/区::'+jsontext['result']['addressComponent']['district'])
    return templist
def get_excel_latlng(startnum,endnum):
    if startnum<=4:
        startnum=4
    addresslist=[]
    sheet_ranges = wb['Sheet1']
    for i in range(startnum-1,endnum):
        lat = str(sheet_ranges['J'+str(i+1)].value)
        lng =str(sheet_ranges['I'+str(i+1)].value)
        addresslist.append(lat+','+lng)
    return addresslist
def write_excel(startnum,endnum,addresslist):
    if startnum<=4:
        startnum=4
    ws = wb['Sheet1']
    for i in range(startnum - 1, endnum):
        num = i-startnum+1
        province,city,district = addresslist[num].split(',')
        ws["L"+str(i+1)] = province
        ws["M"+str(i+1)] = city
        ws["N" + str(i + 1)] = district
    wb.save(file_name)

for i in range(1,71):
    wb = load_workbook(filename=file_name)
    num1 = i * 853 - 853
    num2 = i * 853
    print(i)
    print(str(num1)+'-'+str(num2))
    write_excel(num1,num2,get_address(get_excel_latlng(num1,num2)))